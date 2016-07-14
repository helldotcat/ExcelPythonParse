#! python

import urllib, sys, bs4, socket
from openpyxl import load_workbook
from openpyxl.compat import range

def domainInfoRequest(domain):

    i = 0
    addr = ""
    rank = ""
    country_rank = ""

    domain = domain.lstrip()
    domain = domain.splitlines() #remove \n
    domain = ''.join(domain)
    domain = domain.rstrip()
    domain = domain.lower()

    #if empty string -> leave
    if domain == "":
        result = domain+';'+addr+';'+rank+';'+country_rank
        return result

    else:
        #Contact to Alexa API
        alexa_response = urllib.urlopen("http://data.alexa.com/data?cli=10&dat=s&url="+ domain).read()
        rank_result = bs4.BeautifulSoup(alexa_response, "xml")

        #Get IP by hostname
        try:
            addr = socket.gethostbyname(domain)
        except:
            addr = ""

        #Get rank from response
        try:
            rank = rank_result.find("POPULARITY")['TEXT']
            country_rank = rank_result.find("COUNTRY")['RANK']

        except:
            rank = ""
            country_rank = ""
        finally:
            result = domain+';'+addr+';'+rank+';'+country_rank
    return result

filename_domains = sys.argv[1]
filename_domains = filename_domains.rstrip(".xlsx")
filename_domains = filename_domains+"_ready.xlsx"

wb = load_workbook(sys.argv[1], read_only=False)
#wb = load_workbook(filename='original2.xlsx', read_only=False)
ws = wb.active

#Get table size
row_count = ws.max_row
column_count = ws.max_column

i = 0
progress = 100./row_count

domain_column = 0
ip_column = 0
alexa_rank_column = 0
country_rank_column = 0

for column in range (1, column_count):
    if ws.cell(column=column, row=1).value == "Domain":
        domain_column = column
    elif ws.cell(column=column, row=1).value == "IP":
        ip_column = column
    elif ws.cell(column=column, row=1).value == "Rating":
        alexa_rank_column = column
    elif ws.cell(column=column, row=1).value == "Alexa Rank":
        country_rank_column = column

if domain_column != 0 and ip_column != 0 and alexa_rank_column != 0 and country_rank_column != 0:
    for row in range (2, row_count):
        i += 1
        #Read domain on every row
        domain_name = ws.cell(column=domain_column, row=row).value

        #Check if unicode
        if isinstance(domain_name, unicode):
            domain_name = domain_name.encode('punycode')

            domain_name_len = len(domain_name)-1
            if domain_name[domain_name_len] == '-':
                domain_name = domain_name.rstrip('-')

            #domain info request
            domain_info = domainInfoRequest(domain_name)
            domain_info = domain_info.splitlines() #remove \n
            domain_info = ''.join(domain_info)

            domain_info = domain_info.encode('punycode')
            domain_info_len = len(domain_info)-1
            if domain_info[domain_info_len] == '-':
                domain_info = domain_info.rstrip('-')

            print str(i)+' '+domain_info
            domain_info = domain_info.split(';')

            addr = domain_info[1]
            rank = domain_info[2]
            country_rank = domain_info[3]

            #If IP is not set then write IP
            if not ws.cell(row = row, column = ip_column).value:
                ws.cell(row = row, column = ip_column).value = addr
            ws.cell(row = row, column = alexa_rank_column).value = rank
            ws.cell(row = row, column = country_rank_column).value=country_rank

        counter = i*progress
        sys.stdout.write("\r%3f%% " % counter)

sys.stdout.write("\r%d%%" % 100)
wb.save(filename_domains)


